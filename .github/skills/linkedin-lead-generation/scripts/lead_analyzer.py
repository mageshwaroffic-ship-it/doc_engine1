#!/usr/bin/env python3
"""
LinkedIn Lead Generation Analyzer

Processes LinkedIn posts from CSV and generates lead scores.
Outputs results to Excel (.xlsx) for CRM import and team sharing.

Usage:
    python lead_analyzer.py --input posts.csv --output leads.xlsx
    python lead_analyzer.py --input posts.csv --filter HIGH --min-score 7

Requirements:
    pip install pandas openpyxl
"""

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


# Keywords for lead scoring
PAIN_WORDS = {
    'slow', 'manual', 'time consuming', 'time-consuming', 'tedious', 'repetitive',
    'inefficient', 'error prone', 'error-prone', 'wasting time', 'draining',
    'bottleneck', 'struggling', 'difficult', 'challenging', 'overwhelmed',
    'drowning in', 'stuck with', 'losing time', 'manual data entry', 'copy/paste',
    'spreadsheet', 'excel hell', 'duplicate', 'inconsistency', 'broken workflow',
    'gaps', 'missing automation', 'legacy', 'outdated', 'unreliable'
}

BUYER_WORDS = {
    'looking for', 'need', 'seeking', 'in market for', 'searching for',
    'hunting for', 'require', 'must have', 'urgent', 'asap', 'immediately',
    'this week', 'this month', 'this quarter', 'deadline', 'budget approved',
    'allocated budget', 'have budget', 'funded', 'investment', 'willing to pay',
    'owned by', 'i manage', 'i lead', 'my team', 'we need', 'we\'re looking',
    'we\'re hiring', 'our company', 'decision maker', 'decision-maker'
}

SPAM_WORDS = {
    'follow me', 'link in bio', 'check my profile', 'dm for details',
    'meme', 'joke', 'funny', 'repost', 'congratulations', 'thanks for',
    'opportunity', 'just sharing', 'learning journey', 'hiring', 'available',
    'freelancer for hire', 'open to opportunities'
}

HIGH_INTENT_KEYWORDS = {
    'looking for developer', 'need website developer', 'freelance developer required',
    'need developer', 'hiring developer', 'need automation', 'looking for automation',
    'automation engineer', 'need help', 'hiring', 'seeking', 'in market for'
}


class LinkedInLeadAnalyzer:
    """Analyzes LinkedIn posts for lead generation scoring."""
    
    def __init__(self, min_score: int = 5, min_confidence: str = 'LOW'):
        self.min_score = min_score
        self.min_confidence = min_confidence
        self.results = []
    
    def classify_intent(self, post_text: str) -> Tuple[str, float]:
        """
        Classify post intent as HIGH, MEDIUM, or LOW.
        Returns: (intent, confidence_score 0-1)
        """
        text_lower = post_text.lower()
        
        # Check for spam first
        for spam_word in SPAM_WORDS:
            if spam_word in text_lower:
                return 'LOW', 0.8
        
        # Check HIGH intent signals
        high_intent_count = sum(1 for keyword in HIGH_INTENT_KEYWORDS 
                               if keyword in text_lower)
        if high_intent_count >= 1:
            return 'HIGH', 0.9
        
        # Check for MEDIUM intent (problem mention)
        pain_count = sum(1 for word in PAIN_WORDS if word in text_lower)
        buyer_count = sum(1 for word in BUYER_WORDS if word in text_lower)
        
        if pain_count >= 1 or buyer_count >= 1:
            confidence = 0.6 + (min(pain_count + buyer_count, 5) * 0.08)
            return 'MEDIUM', confidence
        
        return 'LOW', 0.5
    
    def extract_problems(self, post_text: str) -> List[str]:
        """Extract key problems mentioned in the post."""
        sentences = re.split(r'[.!?]+', post_text)
        problems = []
        
        for sentence in sentences:
            sentence_lower = sentence.lower()
            # Find sentences with pain words
            for pain_word in PAIN_WORDS:
                if pain_word in sentence_lower and len(sentence.strip()) > 10:
                    problems.append(sentence.strip()[:100])  # First 100 chars
                    break
        
        return problems if problems else ["Problem not explicitly stated"]
    
    def extract_keywords(self, post_text: str) -> List[str]:
        """Extract relevant keywords from post."""
        text_lower = post_text.lower()
        keywords = []
        
        # Collect pain keywords
        for word in PAIN_WORDS:
            if word in text_lower:
                keywords.append(word)
        
        # Collect buyer keywords
        for word in BUYER_WORDS:
            if word in text_lower:
                keywords.append(word)
        
        return list(set(keywords))[:10]  # Unique, max 10
    
    def score_urgency(self, post_text: str) -> str:
        """Determine urgency level based on language."""
        text_lower = post_text.lower()
        
        high_urgency = {'asap', 'urgent', 'immediately', 'deadline', 'critical'}
        medium_urgency = {'soon', 'next week', 'this month', 'this quarter'}
        
        if any(word in text_lower for word in high_urgency):
            return 'High'
        elif any(word in text_lower for word in medium_urgency):
            return 'Medium'
        else:
            return 'Low'
    
    def score_lead(self, post_text: str, intent: str) -> int:
        """
        Score lead on 1-10 scale.
        
        Scoring:
        - Intent: HIGH=3pts, MEDIUM=1pt, LOW=0pts
        - Pain words: +1pt each (max 5pts)
        - Buyer words: +2pts each (max 4pts)
        - Business relevance: +2pts
        """
        score = 0
        text_lower = post_text.lower()
        
        # Intent score
        if intent == 'HIGH':
            score += 3
        elif intent == 'MEDIUM':
            score += 1
        
        # Pain word score
        pain_count = sum(1 for word in PAIN_WORDS if word in text_lower)
        score += min(pain_count, 5)
        
        # Buyer word score
        buyer_count = sum(1 for word in BUYER_WORDS if word in text_lower)
        score += min(buyer_count * 2, 4)
        
        # Business relevance (software/automation/CRM context)
        business_keywords = {'software', 'automation', 'crm', 'process', 'workflow',
                           'system', 'tool', 'platform', 'business', 'company'}
        if any(keyword in text_lower for keyword in business_keywords):
            score += 2
        
        return min(score, 10)  # Cap at 10
    
    def suggest_service(self, post_text: str) -> str:
        """Suggest relevant service category."""
        text_lower = post_text.lower()
        
        if any(word in text_lower for word in ['crm', 'sales', 'pipeline', 'lead']):
            return 'CRM/Sales Automation'
        elif any(word in text_lower for word in ['website', 'app', 'development', 'developer']):
            return 'Software Development'
        elif any(word in text_lower for word in ['ai', 'chatbot', 'machine learning']):
            return 'AI/ML Automation'
        elif any(word in text_lower for word in ['zapier', 'make', 'integration', 'api']):
            return 'Workflow Integration'
        else:
            return 'Business Process Automation'
    
    def generate_reply(self, author: str, problem: str, service: str) -> str:
        """Generate personalized reply (simplified, actual LLM would be better)."""
        if not problem or problem == "Problem not explicitly stated":
            problem = "your workflow challenges"
        
        templates = [
            f"Hi {author}, saw your post about {problem.lower()}. We help teams solve this with {service.lower()}. Happy to discuss if helpful 👍",
            f"Great question! We specialize in {service.lower()} and have helped teams tackle similar {problem.lower()}. Would love to share ideas 💡",
            f"Totally understand the pain. We work with companies on exactly this—{problem.lower()}. Let me know if you'd like to explore solutions 👍"
        ]
        
        return templates[0][:160]  # Max 160 chars
    
    def analyze_post(self, author: str, post_text: str, post_url: str = "", 
                    date: str = "") -> Dict:
        """Analyze a single LinkedIn post."""
        if not post_text or len(post_text.strip()) < 10:
            return None
        
        intent, confidence = self.classify_intent(post_text)
        
        if intent == 'LOW' and confidence > 0.7:
            return None  # Skip spam/low intent with high confidence
        
        problems = self.extract_problems(post_text)
        primary_problem = problems[0] if problems else "Workflow optimization needed"
        keywords = self.extract_keywords(post_text)
        urgency = self.score_urgency(post_text)
        lead_score = self.score_lead(post_text, intent)
        service = self.suggest_service(post_text)
        reply = self.generate_reply(author, primary_problem, service)
        
        return {
            'author': author,
            'post': post_text[:200],  # Truncate for readability
            'post_url': post_url,
            'date': date,
            'intent': intent,
            'problem': primary_problem,
            'urgency': urgency,
            'keywords': ', '.join(keywords[:5]),  # Top 5 keywords
            'lead_score': lead_score,
            'suggested_service': service,
            'reply': reply,
            'confidence': f"{confidence*100:.0f}%"
        }
    
    def process_csv(self, csv_path: str) -> pd.DataFrame:
        """Process CSV file of LinkedIn posts."""
        df = pd.read_csv(csv_path)
        
        # Expect columns: author, post, url (optional), date (optional)
        required_cols = ['author', 'post']
        missing = [col for col in required_cols if col not in df.columns]
        
        if missing:
            raise ValueError(f"CSV missing required columns: {missing}")
        
        results = []
        for _, row in df.iterrows():
            result = self.analyze_post(
                author=row.get('author', 'Unknown'),
                post_text=row.get('post', ''),
                post_url=row.get('url', ''),
                date=row.get('date', '')
            )
            if result and result['lead_score'] >= self.min_score:
                results.append(result)
        
        return pd.DataFrame(results)
    
    def export_excel(self, df: pd.DataFrame, output_path: str):
        """Export results to formatted Excel file with clickable LinkedIn links."""
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            df.to_csv(output_path.replace('.xlsx', '.csv'), index=False)
            print(f"Exported to CSV: {output_path.replace('.xlsx', '.csv')}")
            return
        
        # Reorder columns: put URL earlier for visibility
        column_order = ['author', 'post_url', 'post', 'intent', 'problem', 'urgency', 
                       'keywords', 'lead_score', 'suggested_service', 'reply', 'confidence', 'date']
        # Only include columns that exist in dataframe
        column_order = [col for col in column_order if col in df.columns]
        df_export = df[column_order]
        
        # Create Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Leads', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Leads']
            
            # Add hyperlinks to post_url column if it exists
            from openpyxl.utils import get_column_letter
            if 'post_url' in df_export.columns:
                url_col_idx = df_export.columns.get_loc('post_url') + 1  # +1 because Excel is 1-indexed
                for row_idx, url in enumerate(df_export['post_url'], start=2):  # start=2 to skip header
                    if pd.notna(url) and str(url).startswith('http'):
                        cell = worksheet.cell(row=row_idx, column=url_col_idx)
                        cell.value = 'View Post'
                        cell.hyperlink = str(url)
                        cell.font = Font(color='0563C1', underline='single')
            
            # Apply formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            # Set column widths and format
            column_widths = {
                'A': 15,  # author
                'B': 15,  # post_url (clickable link)
                'C': 30,  # post
                'D': 12,  # intent
                'E': 30,  # problem
                'F': 12,  # urgency
                'G': 25,  # keywords
                'H': 12,  # lead_score
                'I': 25,  # service
                'J': 35,  # reply
                'K': 12,  # confidence
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format data rows
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    
                    # Color code intent
                    if cell.column_letter == 'C':  # Intent column
                        if cell.value == 'HIGH':
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        elif cell.value == 'MEDIUM':
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    
                    # Color code score
                    if cell.column_letter == 'G':  # Score column
                        try:
                            score = int(cell.value)
                            if score >= 8:
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            elif score >= 6:
                                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        except:
                            pass
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
        
        print(f"✅ Exported {len(df)} leads to Excel: {output_path}")
    
    def filter_results(self, df: pd.DataFrame, intent_filter: str = None, 
                      min_score: int = None) -> pd.DataFrame:
        """Filter results by intent and score."""
        if intent_filter:
            df = df[df['intent'].str.upper() == intent_filter.upper()]
        
        if min_score:
            df = df[df['lead_score'] >= min_score]
        
        return df.sort_values('lead_score', ascending=False).reset_index(drop=True)


def main():
    parser = argparse.ArgumentParser(
        description='LinkedIn Lead Generation Analyzer',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python lead_analyzer.py --input posts.csv --output leads.xlsx
  python lead_analyzer.py --input posts.csv --output leads.xlsx --filter HIGH
  python lead_analyzer.py --input posts.csv --output leads.xlsx --min-score 7
        """
    )
    
    parser.add_argument('--input', '-i', required=True, help='Input CSV file')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file')
    parser.add_argument('--filter', choices=['HIGH', 'MEDIUM', 'LOW'],
                       help='Filter by intent level')
    parser.add_argument('--min-score', type=int, default=5,
                       help='Minimum lead score (1-10)')
    parser.add_argument('--format', choices=['xlsx', 'csv'], default='xlsx',
                       help='Output format')
    
    args = parser.parse_args()
    
    # Check input file
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ Error: Input file not found: {args.input}")
        return
    
    print(f"📊 Processing LinkedIn posts from: {args.input}")
    
    # Analyze
    analyzer = LinkedInLeadAnalyzer(min_score=args.min_score)
    results_df = analyzer.process_csv(args.input)
    
    if len(results_df) == 0:
        print(f"⚠️  No leads found matching criteria (min_score >= {args.min_score})")
        return
    
    # Filter if requested
    if args.filter:
        results_df = analyzer.filter_results(results_df, intent_filter=args.filter)
        print(f"🔍 Filtered to {len(results_df)} {args.filter} intent leads")
    
    # Export
    output_path = args.output
    if args.format == 'xlsx' or output_path.endswith('.xlsx'):
        analyzer.export_excel(results_df, output_path)
    else:
        results_df.to_csv(output_path, index=False)
        print(f"✅ Exported to CSV: {output_path}")
    
    # Summary
    print(f"\n📈 Summary:")
    print(f"   Total leads: {len(results_df)}")
    print(f"   HIGH intent: {len(results_df[results_df['intent'] == 'HIGH'])}")
    print(f"   MEDIUM intent: {len(results_df[results_df['intent'] == 'MEDIUM'])}")
    print(f"   Avg score: {results_df['lead_score'].mean():.1f}")


if __name__ == '__main__':
    main()
